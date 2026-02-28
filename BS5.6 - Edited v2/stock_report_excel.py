from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def export_stock_report(data, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Report"

    headers = [
        "SR NO", "NAME", "UNIT",
        "OPENING STOCK", "PURCHASE QTY", "SALE QTY",
        "CLOSING STOCK", "SALE PRICE",
        "TOTAL VALUE", "PURCHASE PRICE",
        "TOTAL VALUE", "PROFIT"
    ]

    # ---------- Styles ----------
    header_fill = PatternFill("solid", fgColor="00B0F0")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # ---------- Header Row ----------
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin
        ws.column_dimensions[get_column_letter(col)].width = 18

    # ---------- Data Rows ----------
    for i, item in enumerate(data, start=2):
        ws.cell(i, 1, i - 1)
        ws.cell(i, 2, item["name"])
        ws.cell(i, 3, item["unit"])
        ws.cell(i, 4, item["opening"])
        ws.cell(i, 5, item["purchase_qty"])
        ws.cell(i, 6, item["sale_qty"])
        ws.cell(i, 7, item["closing"])
        ws.cell(i, 8, item["sale_price"])

        sale_total = item["sale_qty"] * item["sale_price"]
        purchase_total = item["sale_qty"] * item["purchase_price"]
        profit = sale_total - purchase_total

        ws.cell(i, 9, sale_total)
        ws.cell(i, 10, item["purchase_price"])
        ws.cell(i, 11, purchase_total)
        ws.cell(i, 12, profit)

        for col in range(1, 13):
            ws.cell(i, col).border = thin
            ws.cell(i, col).alignment = center

    wb.save(filename)
